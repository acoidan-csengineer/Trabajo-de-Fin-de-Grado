
# Importación de librerías

import pandas as pd
import getopt, sys
import json
import os.path
from openpyxl import load_workbook

inputfile = ''
sheetExcel = ''
outputfile = ''

# Get full command-line arguments
full_cmd_arguments = sys.argv

# Keep all but the first
argument_list = full_cmd_arguments[1:]

short_options = "hi:s:o:"
long_options = ["help", "input=", "sheet=", "output="]

try:
    options, arguments = getopt.getopt(argument_list, short_options, long_options)
except getopt.GetoptError:
    print("\nERROR\n")
    print('The syntax of the execution is incorrect, it should be: "python <EXECUTABLE> -i input_file_name.xlsx -s sheet_name -o output_file_name.json"\n')
    sys.exit(2)

for option, argument in options:
    if option == '-h':
        print("\nHELP MODE\n")
        print("\nUsage: python <EXECUTABLE> -i input_file_name.xlsx -s sheet_name -o output_file_name.json [-h\n")
        print("Required arguments:")
        print("\t-i input_file_name.xlsx, --input input_file_name.xlsx      Fichero de entrada Excel")
        print("\t-s sheet_name, --sheet sheet_name      Excel sheet containing the data")
        print("\nOptional arguments:")
        print("\t-o output_file_name.json, --output output_file_name.json      JSON output file")
        print("\t-h, --help            Show this help message and exit\n")
        
        sys.exit()
    elif option in ("-i", "--input"):
        if (os.path.exists(argument)):
            inputfile = argument
        else:
            print("\nERROR\n")
            print("The file passed with the -i or --input option does not exist.\n")
            sys.exit()
    elif option in ("-s", "--sheet"):
        sheetExcel = argument
    elif option in ("-o", "--output"):
        outputfile = argument

# Check if sheet Excel exists
wb = load_workbook(inputfile, read_only=True)
if sheetExcel not in wb.sheetnames:
    print("\nERROR\n")
    print("The sheet passed with the -s or --sheet option does not exist.\n")
    sys.exit()

# Comprobate if user pass inputfile and sheet Excel
if (inputfile == "") or (sheetExcel == ""):
    print("\nERROR\n")
    print("Some command line arguments are missing\n")
    print("Run python <EXECUTABLE> -h or python <EXECUTABLE> --help to find out which arguments to pass on the command line\n")
    sys.exit()

# If user not pass outputfile, outputfile will be data-pois.json
if (outputfile == ""):
    outputfile = "../data/data-pois.json"

input_cols = list(range(1, 8))

# Read xslx file
df = pd.read_excel(inputfile, sheet_name = sheetExcel, header = 0, usecols = input_cols)
df.columns = ["town", "postalCode", "categories", "name", "geolocation", "description", "applications"]

# Change geolocation format GeometryCollection
# for i in range(df.shape[0]):
#     if (df['geolocation'].values[i].startswith('Desde')):
#         index_hasta = df['geolocation'].values[i].find("hasta")
#         df['geolocation'].values[i] = {
#             "type": "LineString",
#             "coordinates": [
#                 [df['geolocation'].values[i][6:index_hasta-1]],
#                 [df['geolocation'].values[i][index_hasta+6:-1]]
#             ]
#         }
        
#     else:
#         df['geolocation'].values[i] = {
#                 "type": "Point",
#                 "coordinates": [df['geolocation'].values[i]]
#             }
#     print(df['geolocation'].values[i])


# Transform dataframe to adaptable format JSON
pois = json.loads(df.to_json(orient = "records", force_ascii = False))

for jsonP in pois:
    for key in list(jsonP):
        if (key == "geolocation"):
            if (jsonP[key].startswith('Desde')):
                jsonP["geolocationLineString"] = jsonP.pop(key)
                index_hasta = jsonP["geolocationLineString"].find("hasta")

                primer_punto = jsonP["geolocationLineString"][6:index_hasta-1]
                index_coma = primer_punto.find(",")
                primer_punto = [primer_punto[index_coma+2:], primer_punto[0: index_coma]]

                segundo_punto = jsonP["geolocationLineString"][index_hasta+6:-1]
                index_coma = segundo_punto.find(",")
                segundo_punto = [segundo_punto[index_coma+2:], segundo_punto[0: index_coma]]

                jsonP["geolocationLineString"] = {
                "type": "LineString",
                    "coordinates": [
                        primer_punto,
                        segundo_punto
                    ]
                }
            else:
                jsonP["geolocationPoint"] = jsonP.pop(key)

                coordenadas = jsonP["geolocationPoint"]
                index_coma = coordenadas.find(",")
                coordenadas = [coordenadas[index_coma+2:], coordenadas[0: index_coma]]

                jsonP["geolocationPoint"] = {
                    "type": "Point",
                    "coordinates": coordenadas
                }

        elif (key == "categories"):
            categories = []

            if (jsonP[key].startswith('Playa') or jsonP[key].startswith('Piscina') or jsonP[key].startswith('Beach') or jsonP[key].startswith('Rampa') or jsonP[key].startswith('Club Náutico')):
                categories.append("SUNANDBEACH")
                jsonP["stayHourNumber"] = 5
                jsonP["averageCost"] = 15
                jsonP["openingDays"] = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
            elif (jsonP[key].startswith('Restaurante') or jsonP[key].startswith('Lonja')):
                categories.append("GASTRONOMY")
                jsonP["stayHourNumber"] = 2
                jsonP["averageCost"] = 30
                jsonP["hourly"] = "12:00 - 23:00"
                jsonP["openingDays"] = ["TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
            elif (jsonP[key].startswith('Faro') or jsonP[key].startswith('Ermita') or jsonP[key].startswith('Elemento') or jsonP[key].startswith('Escultura') or jsonP[key].startswith('Bien de Interés Cultural')):
                categories.append("CULTURE")
                jsonP["stayHourNumber"] = 3
                jsonP["averageCost"] = 10
                jsonP["openingDays"] = []
                jsonP["openingDays"] = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
            elif (jsonP[key].startswith('Sendero')):
                categories.append("NATURE")
                categories.append("RURAL")
                jsonP["stayHourNumber"] = 4
                jsonP["averageCost"] = 0
                jsonP["openingDays"] = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
            elif (jsonP[key].startswith('Puerto') or jsonP[key].startswith('Muelle')):
                categories.append("CULTURE")
                categories.append("SUNANDBEACH")
                jsonP["stayHourNumber"] = 3
                jsonP["averageCost"] = 0
                jsonP["openingDays"] = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
            elif (jsonP[key].startswith('Paseo')):
                categories.append("SUNANDBEACH")
                jsonP["stayHourNumber"] = 3
                jsonP["averageCost"] = 20
                jsonP["openingDays"] = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]

            jsonP[key] = categories


# Write outputfile
with open(outputfile, 'w') as output:
    json.dump(pois, output, indent = 2)
    output.close();

print("\nProgram executed and " + outputfile + " file written\n")